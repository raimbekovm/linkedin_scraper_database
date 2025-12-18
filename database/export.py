"""
Экспорт данных из БД в различные форматы (JSON, CSV, Excel)
Миграция данных из Excel в БД
"""
import json
import csv
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from database.models import get_db_manager, Person
from database.operations import ProfileManager


class DataExporter:
    """Экспорт данных из БД"""

    def __init__(self):
        self.db = get_db_manager()
        self.pm = ProfileManager()

    def export_to_json(self, filename='export.json', active_only=True):
        """Экспорт всех профилей в JSON"""
        profiles = self.pm.get_all_profiles(active_only=active_only)

        data = []
        for person in profiles:
            profile_data = {
                'id': person.id,
                'name': person.name,
                'linkedin_url': person.linkedin_url,
                'location': person.location,
                'current_job_title': person.current_job_title,
                'current_company': person.current_company,
                'about': person.about,
                'first_scraped_at': person.first_scraped_at.isoformat() if person.first_scraped_at else None,
                'last_scraped_at': person.last_scraped_at.isoformat() if person.last_scraped_at else None,
                'scrape_count': person.scrape_count,
                'experiences': [
                    {
                        'position_title': exp.position_title,
                        'company_name': exp.company_name,
                        'location': exp.location,
                        'from_date': exp.from_date,
                        'to_date': exp.to_date,
                        'duration': exp.duration,
                        'description': exp.description
                    }
                    for exp in person.experiences
                ],
                'educations': [
                    {
                        'institution_name': edu.institution_name,
                        'degree': edu.degree,
                        'from_date': edu.from_date,
                        'to_date': edu.to_date,
                        'description': edu.description
                    }
                    for edu in person.educations
                ]
            }
            data.append(profile_data)

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"✓ Экспортировано {len(data)} профилей в {filename}")
        return len(data)

    def export_to_csv(self, filename='export.csv', active_only=True):
        """Экспорт профилей в CSV (основная информация)"""
        profiles = self.pm.get_all_profiles(active_only=active_only)

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Заголовки
            writer.writerow([
                'ID', 'Имя', 'Локация', 'Должность', 'Компания',
                'LinkedIn URL', 'Количество мест работы', 'Количество учебных заведений',
                'Первый скрэйп', 'Последний скрэйп', 'Количество скрэйпов'
            ])

            # Данные
            for person in profiles:
                writer.writerow([
                    person.id,
                    person.name,
                    person.location,
                    person.current_job_title,
                    person.current_company,
                    person.linkedin_url,
                    len(person.experiences),
                    len(person.educations),
                    person.first_scraped_at.strftime('%Y-%m-%d %H:%M:%S') if person.first_scraped_at else '',
                    person.last_scraped_at.strftime('%Y-%m-%d %H:%M:%S') if person.last_scraped_at else '',
                    person.scrape_count
                ])

        print(f"✓ Экспортировано {len(profiles)} профилей в {filename}")
        return len(profiles)

    def export_to_excel(self, filename='export.xlsx', active_only=True):
        """Экспорт профилей в Excel с использованием pandas"""
        profiles = self.pm.get_all_profiles(active_only=active_only)

        # Основная информация
        main_data = []
        for person in profiles:
            main_data.append({
                'ID': person.id,
                'Имя': person.name,
                'Локация': person.location,
                'Должность': person.current_job_title,
                'Компания': person.current_company,
                'О себе': person.about[:100] + '...' if person.about and len(person.about) > 100 else person.about,
                'LinkedIn URL': person.linkedin_url,
                'Первый скрэйп': person.first_scraped_at,
                'Последний скрэйп': person.last_scraped_at,
                'Количество скрэйпов': person.scrape_count
            })

        # Опыт работы
        exp_data = []
        for person in profiles:
            for exp in person.experiences:
                exp_data.append({
                    'ID профиля': person.id,
                    'Имя': person.name,
                    'Должность': exp.position_title,
                    'Компания': exp.company_name,
                    'Локация': exp.location,
                    'Начало': exp.from_date,
                    'Конец': exp.to_date,
                    'Длительность': exp.duration
                })

        # Образование
        edu_data = []
        for person in profiles:
            for edu in person.educations:
                edu_data.append({
                    'ID профиля': person.id,
                    'Имя': person.name,
                    'Учебное заведение': edu.institution_name,
                    'Степень': edu.degree,
                    'Начало': edu.from_date,
                    'Конец': edu.to_date
                })

        # Создаем Excel с несколькими листами
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            pd.DataFrame(main_data).to_excel(writer, sheet_name='Профили', index=False)
            pd.DataFrame(exp_data).to_excel(writer, sheet_name='Опыт работы', index=False)
            pd.DataFrame(edu_data).to_excel(writer, sheet_name='Образование', index=False)

        print(f"✓ Экспортировано {len(profiles)} профилей в {filename}")
        print(f"  - Профили: {len(main_data)}")
        print(f"  - Опыт работы: {len(exp_data)}")
        print(f"  - Образование: {len(edu_data)}")
        return len(profiles)


class DataMigrator:
    """Миграция данных из Excel в БД"""

    def __init__(self):
        self.pm = ProfileManager()

    def migrate_from_excel(self, excel_file='data/linkedin_profiles.xlsx'):
        """Мигрировать данные из существующего Excel файла в БД"""
        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            migrated = 0
            errors = 0

            print(f"\nНачинаем миграцию из {excel_file}...")

            # Пропускаем заголовок
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    # Парсим данные из строки
                    name = row[0]
                    location = row[1]
                    job_title = row[2]
                    company = row[3]
                    about = row[4]
                    linkedin_url = row[6] if len(row) > 6 else None
                    experiences_text = row[7] if len(row) > 7 else None
                    educations_text = row[8] if len(row) > 8 else None

                    if not linkedin_url:
                        print(f"  ⚠ Пропущен профиль {name}: нет LinkedIn URL")
                        continue

                    # Парсим опыт работы
                    experiences = []
                    if experiences_text:
                        for exp_line in experiences_text.split('\n'):
                            exp_line = exp_line.strip()
                            if exp_line:
                                # Простой парсинг: "Должность в Компания (даты)"
                                parts = exp_line.split(' в ')
                                if len(parts) >= 2:
                                    position = parts[0]
                                    rest = ' в '.join(parts[1:])
                                    company_part = rest.split('(')[0].strip() if '(' in rest else rest
                                    dates = rest.split('(')[1].split(')')[0] if '(' in rest else None

                                    from_date = to_date = duration = None
                                    if dates and ' - ' in dates:
                                        date_parts = dates.split(' - ')
                                        from_date = date_parts[0].strip()
                                        to_date = date_parts[1].strip().split(',')[0] if len(date_parts) > 1 else None
                                        if len(date_parts) > 1 and ',' in date_parts[1]:
                                            duration = date_parts[1].split(',')[1].strip()

                                    experiences.append({
                                        'position_title': position,
                                        'institution_name': company_part,
                                        'from_date': from_date,
                                        'to_date': to_date,
                                        'duration': duration,
                                        'location': None,
                                        'description': None
                                    })

                    # Парсим образование
                    educations = []
                    if educations_text:
                        for edu_line in educations_text.split('\n'):
                            edu_line = edu_line.strip()
                            if edu_line:
                                # Простой парсинг: "Университет - Степень (даты)"
                                parts = edu_line.split(' - ')
                                institution = parts[0].strip()
                                degree = parts[1].split('(')[0].strip() if len(parts) > 1 and '(' in parts[1] else (parts[1] if len(parts) > 1 else None)
                                dates = parts[1].split('(')[1].split(')')[0] if len(parts) > 1 and '(' in parts[1] else None

                                from_date = to_date = None
                                if dates and ' - ' in dates:
                                    date_parts = dates.split(' - ')
                                    from_date = date_parts[0].strip()
                                    to_date = date_parts[1].strip() if len(date_parts) > 1 else None

                                educations.append({
                                    'institution_name': institution,
                                    'degree': degree,
                                    'from_date': from_date,
                                    'to_date': to_date,
                                    'description': None
                                })

                    # Сохраняем в БД
                    profile_data = {
                        'linkedin_url': linkedin_url,
                        'name': name,
                        'location': location,
                        'job_title': job_title,
                        'company': company,
                        'about': about,
                        'experiences': experiences,
                        'educations': educations
                    }

                    self.pm.save_profile(profile_data, track_changes=False)
                    migrated += 1

                except Exception as e:
                    errors += 1
                    print(f"  ✗ Ошибка при обработке строки: {e}")

            print(f"\n✓ Миграция завершена:")
            print(f"  - Успешно: {migrated}")
            print(f"  - Ошибок: {errors}")

            return migrated, errors

        except Exception as e:
            print(f"✗ Ошибка при миграции: {e}")
            raise


if __name__ == "__main__":
    # Тестирование
    exporter = DataExporter()
    migrator = DataMigrator()

    print("Модуль экспорта и миграции готов к работе!")

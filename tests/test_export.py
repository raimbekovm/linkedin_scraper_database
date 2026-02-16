"""Tests for DataExporter (JSON, CSV, Excel)."""

import json
import csv
import os

from database.export import DataExporter


class TestExportJSON:

    def test_export_json_structure(self, populated_db, tmp_path):
        exporter = DataExporter()
        filepath = str(tmp_path / 'out.json')
        count = exporter.export_to_json(filepath)

        assert count == 3
        with open(filepath, encoding='utf-8') as f:
            data = json.load(f)

        assert len(data) == 3
        profile = data[0]
        assert 'name' in profile
        assert 'linkedin_url' in profile
        assert 'experiences' in profile
        assert 'educations' in profile
        assert isinstance(profile['experiences'], list)

    def test_export_json_active_only(self, populated_db, pm, tmp_path):
        pm.delete_profile(populated_db[0].id, soft_delete=True)

        exporter = DataExporter()
        filepath = str(tmp_path / 'active.json')
        count = exporter.export_to_json(filepath, active_only=True)
        assert count == 2

    def test_export_json_empty_db(self, db_manager, tmp_path):
        exporter = DataExporter()
        filepath = str(tmp_path / 'empty.json')
        count = exporter.export_to_json(filepath)
        assert count == 0

        with open(filepath, encoding='utf-8') as f:
            data = json.load(f)
        assert data == []


class TestExportCSV:

    def test_export_csv_structure(self, populated_db, tmp_path):
        exporter = DataExporter()
        filepath = str(tmp_path / 'out.csv')
        count = exporter.export_to_csv(filepath)

        assert count == 3
        with open(filepath, encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)

        header = rows[0]
        assert 'ID' in header
        assert len(rows) == 4  # header + 3 data rows

    def test_export_csv_active_only(self, populated_db, pm, tmp_path):
        pm.delete_profile(populated_db[0].id, soft_delete=True)

        exporter = DataExporter()
        filepath = str(tmp_path / 'active.csv')
        count = exporter.export_to_csv(filepath, active_only=True)
        assert count == 2


class TestExportExcel:

    def test_export_excel_sheets(self, populated_db, tmp_path):
        exporter = DataExporter()
        filepath = str(tmp_path / 'out.xlsx')
        count = exporter.export_to_excel(filepath)

        assert count == 3
        assert os.path.exists(filepath)

        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        assert 'Profiles' in wb.sheetnames
        assert 'Experience' in wb.sheetnames
        assert 'Education' in wb.sheetnames
        wb.close()

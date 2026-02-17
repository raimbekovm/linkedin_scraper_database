"""
Search LinkedIn for alumni by name and scrape their profiles.

Reads names from an Excel or CSV file, searches LinkedIn for each person,
and scrapes the found profiles into the database.

Usage:
    python scripts/search_and_scrape.py data/alumni.csv
    python scripts/search_and_scrape.py data/alumni.xlsx
"""

import csv
import time
import sys
import os
import logging
import argparse

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from linkedin_scraper.person_search import PersonSearch
from database.operations import ProfileManager
from database.models import get_db_manager
from scripts.scrape_to_database import (
    create_driver,
    authenticate,
    scrape_profile_to_db,
)

logger = logging.getLogger(__name__)

DELAY_BETWEEN_SEARCHES = 5
DELAY_BETWEEN_SCRAPES = 3


def read_names_from_csv(filepath: str) -> list[str]:
    """
    Read person names from a CSV file.

    Expects a column named 'name' (case-insensitive) or uses the first column.
    """
    names = []
    with open(filepath, encoding='utf-8') as f:
        reader = csv.DictReader(f)

        # Find the name column (case-insensitive)
        name_col = None
        if reader.fieldnames:
            for col in reader.fieldnames:
                if col.strip().lower() == 'name':
                    name_col = col
                    break
            if name_col is None:
                name_col = reader.fieldnames[0]

        for row in reader:
            value = row.get(name_col, '').strip()
            if value:
                names.append(value)

    return names


def read_names_from_excel(filepath: str) -> list[str]:
    """
    Read person names from an Excel file.

    Expects a column named 'name' (case-insensitive) in the header row,
    or uses the first column.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    names = []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return names

    # Detect name column from header
    header = [str(c).strip().lower() if c else '' for c in rows[0]]
    name_idx = 0
    for idx, col in enumerate(header):
        if col == 'name':
            name_idx = idx
            break

    for row in rows[1:]:
        if row and len(row) > name_idx and row[name_idx]:
            value = str(row[name_idx]).strip()
            if value:
                names.append(value)

    wb.close()
    return names


def read_names(filepath: str) -> list[str]:
    """Read names from CSV or Excel file based on extension."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        return read_names_from_csv(filepath)
    elif ext in ('.xlsx', '.xls'):
        return read_names_from_excel(filepath)
    else:
        raise ValueError(f"Unsupported file format: {ext}. Use .csv or .xlsx")


def main():
    parser = argparse.ArgumentParser(
        description="Search LinkedIn for alumni by name and scrape their profiles."
    )
    parser.add_argument(
        'file',
        help="Path to CSV or Excel file with alumni names",
    )
    parser.add_argument(
        '--limit', type=int, default=1,
        help="How many search results to scrape per name (default: 1)",
    )
    parser.add_argument(
        '--skip', type=int, default=0,
        help="Skip first N names (default: 0)",
    )
    parser.add_argument(
        '--max-names', type=int, default=0,
        help="Max number of names to process after skip (default: 0 = all)",
    )
    parser.add_argument(
        '--school-id', type=str, default=None,
        help="LinkedIn school ID to filter search (e.g. 316375 for AUCA)",
    )
    parser.add_argument(
        '--search-delay', type=float, default=DELAY_BETWEEN_SEARCHES,
        help=f"Seconds to wait between searches (default: {DELAY_BETWEEN_SEARCHES})",
    )
    parser.add_argument(
        '--scrape-delay', type=float, default=DELAY_BETWEEN_SCRAPES,
        help=f"Seconds to wait between scrapes (default: {DELAY_BETWEEN_SCRAPES})",
    )
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"ERROR: File not found: {args.file}")
        sys.exit(1)

    # Read names
    names = read_names(args.file)
    if not names:
        print("ERROR: No names found in the file.")
        sys.exit(1)

    if args.skip > 0:
        names = names[args.skip:]
    if args.max_names > 0:
        names = names[:args.max_names]

    print("=" * 60)
    print("LINKEDIN ALUMNI SEARCH & SCRAPE")
    print("=" * 60)
    print(f"File: {args.file}")
    print(f"Names to process: {len(names)}")
    print(f"Results per name: {args.limit}")
    if args.school_id:
        print(f"School filter: {args.school_id}")

    # Initialize database
    db = get_db_manager()
    db.create_all_tables()
    pm = ProfileManager()

    stats = db.get_stats()
    print(f"\nCurrent DB: {stats['total_persons']} profiles")

    # Setup browser and login
    driver = create_driver()

    try:
        authenticate(driver)
        searcher = PersonSearch(driver)

        # Counters
        searched = 0
        found = 0
        scraped = 0
        skipped = 0
        not_found = 0
        failed = 0

        print("\n" + "=" * 60)
        print(f"SEARCHING {len(names)} NAMES")
        print("=" * 60)

        for idx, name in enumerate(names, 1):
            print(f"\n[{idx}/{len(names)}] Searching: {name}...", end=" ")
            searched += 1

            results = searcher.search(name, limit=args.limit, school_id=args.school_id)

            if not results:
                print("NOT FOUND")
                not_found += 1
                if idx < len(names):
                    time.sleep(args.search_delay)
                continue

            for result in results:
                url = result.linkedin_url
                print(f"\n  Found: {url}")
                found += 1

                # Check if already in DB
                existing = pm.get_profile_by_url(url)
                if existing:
                    print(f"  SKIP: Already in DB (ID: {existing.id}, scrape_count: {existing.scrape_count})")
                    skipped += 1
                    continue

                # Scrape the profile
                print(f"  Scraping...", end=" ")
                saved = scrape_profile_to_db(url, driver, pm, track_changes=False)

                if saved:
                    scraped += 1
                else:
                    failed += 1

                time.sleep(args.scrape_delay)

            if idx < len(names):
                time.sleep(args.search_delay)

        # Summary
        print("\n" + "=" * 60)
        print("SUMMARY")
        print("=" * 60)
        print(f"  Names searched:  {searched}")
        print(f"  Profiles found:  {found}")
        print(f"  Scraped (new):   {scraped}")
        print(f"  Skipped (in DB): {skipped}")
        print(f"  Not found:       {not_found}")
        print(f"  Failed:          {failed}")

        stats = db.get_stats()
        print(f"\nUpdated DB: {stats['total_persons']} profiles, "
              f"{stats['total_experiences']} experiences, "
              f"{stats['total_educations']} educations")

        print("\n" + "=" * 60)
        print("Browser will close in 5 seconds...")
        time.sleep(5)

    except KeyboardInterrupt:
        print("\n\nInterrupted by user.")
    except Exception as e:
        print(f"\nCritical error: {e}")
        logger.exception("Critical error in search_and_scrape")
    finally:
        driver.quit()
        print("Browser closed.")


if __name__ == "__main__":
    main()
